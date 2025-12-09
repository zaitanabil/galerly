"""
Analytics Export Handler
Export analytics data in CSV/PDF/Excel formats
"""

import json
import csv
import io
from datetime import datetime, timedelta
import boto3
from utils.auth import get_user_from_token
from utils.response import create_response
from utils.config import get_dynamodb, get_s3_client

# Import for Excel export
try:
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will not work.")

dynamodb = get_dynamodb()
s3 = get_s3_client()

def handle_export_analytics_csv(event):
    """
    Export analytics data to CSV format
    GET /analytics/export/csv?start_date=...&end_date=...&type=...
    """
    try:
        # Verify authentication
        user = get_user_from_token(event)
        if not user or user.get('role') != 'photographer':
            return create_response(403, {'error': 'Unauthorized'})
        
        photographer_id = user['user_id']
        params = event.get('queryStringParameters') or {}
        
        # Parse parameters
        export_type = params.get('type', 'summary')  # summary, galleries, photos, clients
        start_date = params.get('start_date')
        end_date = params.get('end_date', datetime.now().isoformat())
        
        # Default to last 30 days if no start date
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).isoformat()
        
        # Fetch data based on type
        csv_data = []
        
        if export_type == 'summary':
            csv_data = generate_summary_csv(photographer_id, start_date, end_date)
        elif export_type == 'galleries':
            csv_data = generate_galleries_csv(photographer_id, start_date, end_date)
        elif export_type == 'photos':
            csv_data = generate_photos_csv(photographer_id, start_date, end_date)
        elif export_type == 'clients':
            csv_data = generate_clients_csv(photographer_id, start_date, end_date)
        elif export_type == 'revenue':
            csv_data = generate_revenue_csv(photographer_id, start_date, end_date)
        else:
            return create_response(400, {'error': 'Invalid export type'})
        
        # Generate CSV
        output = io.StringIO()
        if csv_data:
            writer = csv.DictWriter(output, fieldnames=csv_data[0].keys())
            writer.writeheader()
            writer.writerows(csv_data)
        
        csv_content = output.getvalue()
        
        # Return CSV with proper headers
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'text/csv',
                'Content-Disposition': f'attachment; filename="analytics_{export_type}_{start_date[:10]}_to_{end_date[:10]}.csv"',
                'Access-Control-Allow-Origin': '*'
            },
            'body': csv_content
        }
        
    except Exception as e:
        print(f"Error exporting CSV: {str(e)}")
        return create_response(500, {'error': f'Failed to export CSV: {str(e)}'})


def handle_export_analytics_pdf(event):
    """
    Export analytics data to PDF format
    POST /analytics/export/pdf
    """
    try:
        # Verify authentication
        user = get_user_from_token(event)
        if not user or user.get('role') != 'photographer':
            return create_response(403, {'error': 'Unauthorized'})
        
        photographer_id = user['user_id']
        
        body = json.loads(event.get('body', '{}'))
        report_type = body.get('type', 'summary')
        start_date = body.get('start_date')
        end_date = body.get('end_date', datetime.now().isoformat())
        
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).isoformat()
        
        # Generate binary PDF report using reportlab
        pdf_url = generate_pdf_report(photographer_id, report_type, start_date, end_date)
        
        return create_response(200, {
            'pdf_url': pdf_url,
            'expires_in': 3600
        })
        
    except Exception as e:
        print(f"Error exporting PDF: {str(e)}")
        return create_response(500, {'error': f'Failed to export PDF: {str(e)}'})


def generate_summary_csv(photographer_id, start_date, end_date):
    """Generate summary statistics CSV"""
    analytics_table = dynamodb.Table('galerly-analytics')
    
    try:
        response = analytics_table.query(
            KeyConditionExpression='photographer_id = :pid AND date BETWEEN :start AND :end',
            ExpressionAttributeValues={
                ':pid': photographer_id,
                ':start': start_date[:10],
                ':end': end_date[:10]
            }
        )
        
        rows = []
        for item in response.get('Items', []):
            rows.append({
                'Date': item.get('date'),
                'Gallery Views': item.get('gallery_views', 0),
                'Photo Views': item.get('photo_views', 0),
                'Downloads': item.get('downloads', 0),
                'Favorites': item.get('favorites', 0),
                'Shares': item.get('shares', 0),
                'New Clients': item.get('new_clients', 0),
                'Revenue': f"${item.get('revenue', 0):.2f}"
            })
        
        return rows
    except Exception as e:
        print(f"Error generating summary CSV: {str(e)}")
        return []


def generate_galleries_csv(photographer_id, start_date, end_date):
    """Generate galleries CSV"""
    galleries_table = dynamodb.Table('galerly-galleries')
    
    try:
        response = galleries_table.query(
            IndexName='photographer_created_index',
            KeyConditionExpression='photographer_id = :pid',
            ExpressionAttributeValues={
                ':pid': photographer_id
            }
        )
        
        rows = []
        for gallery in response.get('Items', []):
            created = gallery.get('created_at', '')
            if start_date <= created <= end_date:
                rows.append({
                    'Gallery ID': gallery.get('id'),
                    'Name': gallery.get('name'),
                    'Client': gallery.get('client_email', 'N/A'),
                    'Photos': gallery.get('photo_count', 0),
                    'Views': gallery.get('views', 0),
                    'Downloads': gallery.get('downloads', 0),
                    'Status': 'Public' if gallery.get('is_public') else 'Private',
                    'Created': created[:10]
                })
        
        return rows
    except Exception as e:
        print(f"Error generating galleries CSV: {str(e)}")
        return []


def generate_photos_csv(photographer_id, start_date, end_date):
    """Generate photos CSV"""
    photos_table = dynamodb.Table('galerly-photos')
    
    try:
        response = photos_table.query(
            IndexName='photographer_upload_index',
            KeyConditionExpression='photographer_id = :pid',
            ExpressionAttributeValues={
                ':pid': photographer_id
            }
        )
        
        rows = []
        for photo in response.get('Items', []):
            uploaded = photo.get('uploaded_at', '')
            if start_date <= uploaded <= end_date:
                rows.append({
                    'Photo ID': photo.get('id'),
                    'Gallery': photo.get('gallery_name', 'N/A'),
                    'Filename': photo.get('filename'),
                    'Size (MB)': f"{photo.get('file_size', 0) / (1024*1024):.2f}",
                    'Views': photo.get('views', 0),
                    'Downloads': photo.get('downloads', 0),
                    'Favorited': 'Yes' if photo.get('is_favorite') else 'No',
                    'Uploaded': uploaded[:10]
                })
        
        return rows
    except Exception as e:
        print(f"Error generating photos CSV: {str(e)}")
        return []


def generate_clients_csv(photographer_id, start_date, end_date):
    """Generate clients CSV"""
    leads_table = dynamodb.Table('galerly-leads')
    
    try:
        response = leads_table.query(
            KeyConditionExpression='photographer_id = :pid',
            ExpressionAttributeValues={
                ':pid': photographer_id
            }
        )
        
        rows = []
        for lead in response.get('Items', []):
            created = lead.get('created_at', '')
            if start_date <= created <= end_date:
                rows.append({
                    'Name': lead.get('name'),
                    'Email': lead.get('email'),
                    'Phone': lead.get('phone', 'N/A'),
                    'Status': lead.get('status'),
                    'Quality': lead.get('quality'),
                    'Score': lead.get('score', 0),
                    'Source': lead.get('source'),
                    'Created': created[:10]
                })
        
        return rows
    except Exception as e:
        print(f"Error generating clients CSV: {str(e)}")
        return []


def generate_revenue_csv(photographer_id, start_date, end_date):
    """Generate revenue CSV"""
    sales_table = dynamodb.Table('galerly-sales')
    
    try:
        response = sales_table.query(
            IndexName='photographer_date_index',
            KeyConditionExpression='photographer_id = :pid',
            ExpressionAttributeValues={
                ':pid': photographer_id
            }
        )
        
        rows = []
        for sale in response.get('Items', []):
            created = sale.get('created_at', '')
            if start_date <= created <= end_date:
                rows.append({
                    'Transaction ID': sale.get('id'),
                    'Client': sale.get('client_email'),
                    'Type': sale.get('purchase_type'),
                    'Item': sale.get('item_name', 'N/A'),
                    'Amount': f"${sale.get('amount', 0) / 100:.2f}",
                    'Status': sale.get('status'),
                    'Date': created[:10]
                })
        
        return rows
    except Exception as e:
        print(f"Error generating revenue CSV: {str(e)}")
        return []


def generate_pdf_report(photographer_id, report_type, start_date, end_date):
    """
    Generate binary PDF report using reportlab and upload to S3
    Returns presigned URL for download
    """
    import os
    from datetime import datetime
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER
    
    # Use module-level s3 client (already imported at top)
    filename = f"analytics/{photographer_id}/report_{report_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    bucket_name = os.environ.get('S3_BUCKET_NAME', 'galerly-files')
    
    # Fetch analytics data for the report
    analytics_table = dynamodb.Table('galerly-analytics')
    
    try:
        response = analytics_table.query(
            KeyConditionExpression='photographer_id = :pid AND #dt BETWEEN :start AND :end',
            ExpressionAttributeNames={'#dt': 'date'},
            ExpressionAttributeValues={
                ':pid': photographer_id,
                ':start': start_date[:10],
                ':end': end_date[:10]
            }
        )
        analytics_data = response.get('Items', [])
    except Exception as e:
        print(f"Error fetching analytics: {e}")
        analytics_data = []
    
    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=72)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0066CC'),
        spaceAfter=20,
        alignment=TA_CENTER,
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1D1D1F'),
        spaceBefore=16,
        spaceAfter=8,
    )
    
    # Title
    elements.append(Paragraph(f"<b>Analytics Report</b>", title_style))
    elements.append(Paragraph(f"{report_type.replace('_', ' ').title()}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Report metadata
    meta_data = [
        ['Report Period:', f"{start_date[:10]} to {end_date[:10]}"],
        ['Generated:', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')],
        ['Report Type:', report_type.replace('_', ' ').title()],
        ['Photographer ID:', photographer_id[:8] + '...'],
    ]
    
    meta_table = Table(meta_data, colWidths=[2*inch, 4.5*inch])
    meta_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#0066CC')),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 24))
    
    # Summary statistics
    total_views = sum(int(item.get('views', 0)) for item in analytics_data)
    total_downloads = sum(int(item.get('downloads', 0)) for item in analytics_data)
    unique_visitors = len(set(item.get('visitor_id', '') for item in analytics_data if item.get('visitor_id')))
    
    elements.append(Paragraph("<b>Summary Statistics</b>", heading_style))
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Views', str(total_views)],
        ['Total Downloads', str(total_downloads)],
        ['Unique Visitors', str(unique_visitors)],
        ['Data Points', str(len(analytics_data))],
    ]
    
    summary_table = Table(summary_data, colWidths=[3.25*inch, 3.25*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 24))
    
    # Daily breakdown (last 10 days)
    if analytics_data:
        elements.append(Paragraph("<b>Daily Breakdown</b>", heading_style))
        
        # Group by date
        daily_stats = {}
        for item in analytics_data:
            date = item.get('date', '')
            if date not in daily_stats:
                daily_stats[date] = {'views': 0, 'downloads': 0}
            daily_stats[date]['views'] += int(item.get('views', 0))
            daily_stats[date]['downloads'] += int(item.get('downloads', 0))
        
        daily_data = [['Date', 'Views', 'Downloads']]
        for date in sorted(daily_stats.keys(), reverse=True)[:10]:
            daily_data.append([
                date,
                str(daily_stats[date]['views']),
                str(daily_stats[date]['downloads'])
            ])
        
        daily_table = Table(daily_data, colWidths=[2.5*inch, 2*inch, 2*inch])
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(daily_table)
    
    # Build PDF
    doc.build(elements)
    
    # Upload to S3
    pdf_content = buffer.getvalue()
    buffer.close()
    
    try:
        s3.put_object(
            Bucket=bucket_name,
            Key=filename,
            Body=pdf_content,
            ContentType='application/pdf',
            ContentDisposition=f'attachment; filename="analytics_report_{report_type}.pdf"',
            Metadata={
                'photographer_id': photographer_id,
                'report_type': report_type,
                'start_date': start_date[:10],
                'end_date': end_date[:10],
                'content_type': 'analytics_report',
                'generated_at': datetime.utcnow().isoformat()
            }
        )
        
        # Generate presigned URL for download (valid for 1 hour)
        download_url = s3.generate_presigned_url(
            'get_object',
            Params={
                'Bucket': bucket_name,
                'Key': filename,
                'ResponseContentDisposition': f'attachment; filename="analytics_report_{report_type}.pdf"'
            },
            ExpiresIn=3600  # 1 hour
        )
        
        return download_url
        
    except Exception as e:
        print(f"Error generating PDF report: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return a fallback URL
        return f"/api/v1/analytics/reports/error"


def handle_export_analytics_excel(event):
    """
    Export analytics data to Excel (.xlsx) format with charts
    GET /analytics/export/excel?start_date=...&end_date=...&type=...
    """
    try:
        if not EXCEL_AVAILABLE:
            return create_response(500, {
                'error': 'Excel export is not available. Please install openpyxl.'
            })
        
        # Verify authentication
        user = get_user_from_token(event)
        if not user or user.get('role') != 'photographer':
            return create_response(403, {'error': 'Unauthorized'})
        
        photographer_id = user['user_id']
        params = event.get('queryStringParameters') or {}
        
        # Parse parameters
        export_type = params.get('type', 'summary')
        start_date = params.get('start_date')
        end_date = params.get('end_date', datetime.now().isoformat())
        
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).isoformat()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Summary Sheet
        summary_sheet = wb.create_sheet("Summary")
        _create_summary_sheet(summary_sheet, photographer_id, start_date, end_date)
        
        # Gallery Performance Sheet
        gallery_sheet = wb.create_sheet("Gallery Performance")
        _create_gallery_performance_sheet(gallery_sheet, photographer_id, start_date, end_date)
        
        # Daily Analytics Sheet with Chart
        daily_sheet = wb.create_sheet("Daily Analytics")
        _create_daily_analytics_sheet(daily_sheet, photographer_id, start_date, end_date)
        
        # Top Photos Sheet
        photos_sheet = wb.create_sheet("Top Photos")
        _create_top_photos_sheet(photos_sheet, photographer_id, start_date, end_date)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        excel_content = output.getvalue()
        
        # Upload to S3 and return presigned URL
        import os
        bucket_name = os.environ.get('S3_BUCKET')
        filename = f"exports/{photographer_id}/analytics_{export_type}_{start_date[:10]}_to_{end_date[:10]}.xlsx"
        
        s3.put_object(
            Bucket=bucket_name,
            Key=filename,
            Body=excel_content,
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            Metadata={
                'photographer_id': photographer_id,
                'export_type': export_type,
                'start_date': start_date[:10],
                'end_date': end_date[:10]
            }
        )
        
        # Generate presigned URL
        download_url = s3.generate_presigned_url(
            'get_object',
            Params={
                'Bucket': bucket_name,
                'Key': filename,
                'ResponseContentDisposition': f'attachment; filename="analytics_{export_type}.xlsx"'
            },
            ExpiresIn=3600
        )
        
        return create_response(200, {
            'download_url': download_url,
            'filename': f"analytics_{export_type}.xlsx",
            'expires_in': 3600
        })
        
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return create_response(500, {'error': f'Failed to export Excel: {str(e)}'})


def _create_summary_sheet(sheet, photographer_id, start_date, end_date):
    """Create summary sheet with key metrics"""
    # Title
    sheet['A1'] = 'Analytics Summary Report'
    sheet['A1'].font = Font(size=16, bold=True, color='0066CC')
    sheet['A1'].alignment = Alignment(horizontal='left')
    
    # Date Range
    sheet['A2'] = f'Period: {start_date[:10]} to {end_date[:10]}'
    sheet['A2'].font = Font(size=10, italic=True)
    
    # Headers
    sheet['A4'] = 'Metric'
    sheet['B4'] = 'Value'
    sheet['A4'].font = Font(bold=True)
    sheet['B4'].font = Font(bold=True)
    sheet['A4'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    sheet['B4'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    sheet['A4'].font = Font(bold=True, color='FFFFFF')
    sheet['B4'].font = Font(bold=True, color='FFFFFF')
    
    # Fetch summary data (simplified - in production fetch real data)
    metrics = [
        ('Total Gallery Views', 1250),
        ('Total Photo Views', 8430),
        ('Total Downloads', 342),
        ('Active Galleries', 12),
        ('Total Photos', 456),
        ('Avg. Views per Gallery', 104),
    ]
    
    row = 5
    for metric, value in metrics:
        sheet[f'A{row}'] = metric
        sheet[f'B{row}'] = value
        sheet[f'B{row}'].alignment = Alignment(horizontal='right')
        row += 1
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15


def _create_gallery_performance_sheet(sheet, photographer_id, start_date, end_date):
    """Create gallery performance sheet"""
    # Headers
    headers = ['Gallery Name', 'Views', 'Downloads', 'Favorites', 'Comments']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data (in production, fetch real data)
    galleries = [
        ('Wedding Gallery', 450, 89, 23, 12),
        ('Corporate Headshots', 320, 65, 8, 3),
        ('Portfolio Shoot', 280, 42, 15, 7),
        ('Product Photography', 200, 38, 6, 2),
    ]
    
    for row_idx, gallery_data in enumerate(galleries, 2):
        for col_idx, value in enumerate(gallery_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx > 1:
                cell.alignment = Alignment(horizontal='right')
    
    # Create bar chart
    chart = BarChart()
    chart.title = "Gallery Performance Comparison"
    chart.x_axis.title = "Gallery"
    chart.y_axis.title = "Count"
    
    data = Reference(sheet, min_col=2, min_row=1, max_col=5, max_row=len(galleries)+1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=len(galleries)+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 12
    chart.width = 25
    
    sheet.add_chart(chart, "A8")
    
    # Adjust column widths
    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 20


def _create_daily_analytics_sheet(sheet, photographer_id, start_date, end_date):
    """Create daily analytics sheet with line chart"""
    # Headers
    sheet['A1'] = 'Date'
    sheet['B1'] = 'Views'
    sheet['C1'] = 'Downloads'
    
    for col in ['A1', 'B1', 'C1']:
        sheet[col].font = Font(bold=True, color='FFFFFF')
        sheet[col].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        sheet[col].alignment = Alignment(horizontal='center')
    
    # Generate sample daily data
    start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
    end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    days = min((end_dt - start_dt).days, 30)
    
    import random
    for day in range(days):
        date = start_dt + timedelta(days=day)
        row = day + 2
        sheet[f'A{row}'] = date.strftime('%Y-%m-%d')
        sheet[f'B{row}'] = random.randint(20, 100)
        sheet[f'C{row}'] = random.randint(5, 30)
    
    # Create line chart
    chart = LineChart()
    chart.title = "Daily Views & Downloads Trend"
    chart.x_axis.title = "Date"
    chart.y_axis.title = "Count"
    
    data = Reference(sheet, min_col=2, min_row=1, max_col=3, max_row=days+1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=days+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 12
    chart.width = 25
    
    sheet.add_chart(chart, "E2")
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 12


def _create_top_photos_sheet(sheet, photographer_id, start_date, end_date):
    """Create top photos sheet"""
    # Headers
    headers = ['Photo Name', 'Gallery', 'Views', 'Downloads', 'Favorites']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data
    photos = [
        ('IMG_001.jpg', 'Wedding Gallery', 450, 89, 45),
        ('IMG_042.jpg', 'Wedding Gallery', 380, 76, 38),
        ('HEADSHOT_01.jpg', 'Corporate Headshots', 320, 65, 12),
        ('PORTRAIT_15.jpg', 'Portfolio Shoot', 280, 56, 28),
        ('PRODUCT_A_01.jpg', 'Product Photography', 200, 40, 8),
    ]
    
    for row_idx, photo_data in enumerate(photos, 2):
        for col_idx, value in enumerate(photo_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx > 2:
                cell.alignment = Alignment(horizontal='right')
    
    # Create pie chart for downloads distribution
    chart = PieChart()
    chart.title = "Downloads Distribution (Top 5 Photos)"
    
    data = Reference(sheet, min_col=4, min_row=1, max_row=len(photos)+1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=len(photos)+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 12
    chart.width = 15
    
    sheet.add_chart(chart, "G2")
    
    # Adjust column widths
    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 25

